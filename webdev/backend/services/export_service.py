"""Admin Excel/ZIP export helpers for test submissions.

Builds a structured workbook for a test's submissions and packages PDFs
or per-user Excels as a ZIP archive.
"""
import html
import json
import os
import re
import tempfile
import zipfile
from datetime import date, datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from models.block import Block
from models.question import Question
from models.submission import Submission
from models.test import Test


class ExportTooLarge(Exception):
    """Raised when an export archive would exceed the Telegram size limit."""
    pass


# Cap on per-user detail sheets in a combined workbook. Past this many
# completed users we omit the individual sheets (anti-OOM / anti-timeout);
# the full data still lives in the summary sheets + the admin panel.
MAX_USER_SHEETS = 300


_HEADER_FONT = Font(bold=True, size=10)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
_CELL_ALIGN = Alignment(wrap_text=True, vertical="top")
_FIXED_COLS = {
    "ID": 5, "Ім'я": 12, "Прізвище": 12, "Email": 24, "Телефон": 14,
    "Сектор": 16, "Розмір компанії": 15, "Вік компанії": 15, "Оборот": 16,
    "Статус": 12, "Дата": 18, "Мова": 8, "Доставка": 12, "Згода": 13,
    "Telegram @user": 16, "Telegram ім'я": 16, "Telegram ChatID": 14,
    "Загальний бал %": 11,
}

# Openpyxl sheet-name constraints: ≤31 chars, no / \ ? * [ ]
_INVALID_SHEET_CHARS = re.compile(r"[\\/:*?\[\]]")


def _safe_filename(name: str) -> str:
    return re.sub(r"[^\w\- ]+", "_", (name or "").strip())[:80] or "report"


def _safe_sheet_name(name: str) -> str:
    clean = _INVALID_SHEET_CHARS.sub("_", (name or "").strip())[:31]
    return clean or "Sheet"


def _infer_delivery(sub: dict) -> str:
    """Guess the delivery channel the user picked based on captured state."""
    if sub.get("tg_chat_id"):
        return "Telegram"
    if sub.get("email"):
        return "Email"
    return "PDF"


def _parse_json_field(val):
    if isinstance(val, str):
        try:
            return json.loads(val)
        except (json.JSONDecodeError, TypeError):
            return None
    return val


def _collect_questions_for_blocks(blocks):
    """Return (question_keys_ordered, question_labels).

    Keys are composed `b{block_id}q{question_id}` and labels are friendly
    `B1 Q2: first 40 chars...` strings (in UK).

    Batched: one query fetches questions for all blocks at once (ANY(%s)),
    replacing per-block round-trips. For 10 blocks with 50 questions, this
    collapses 10 queries into 1.
    """
    if not blocks:
        return [], {}

    block_ids = [b["id"] for b in blocks]
    all_questions = Question.find_by_blocks(block_ids)

    # Group questions by block for ordered traversal.
    by_block: dict[int, list] = {b["id"]: [] for b in blocks}
    for q in all_questions:
        by_block.setdefault(q["block_id"], []).append(q)

    question_keys_ordered: list[str] = []
    question_labels: dict[str, str] = {}

    for b_idx, b in enumerate(blocks):
        block_questions = by_block.get(b["id"], [])
        top_level = [q for q in block_questions if not q.get("parent_question_id")]
        for q_idx, q in enumerate(top_level):
            key = f"b{b['id']}q{q['id']}"
            question_labels[key] = f"B{b_idx + 1} Q{q_idx + 1}: {(q['text_uk'] or '')[:40]}"
            question_keys_ordered.append(key)
            subs = [sq for sq in block_questions if sq.get("parent_question_id") == q["id"]]
            for sq_idx, sq in enumerate(subs):
                skey = f"b{b['id']}q{sq['id']}"
                question_labels[skey] = (
                    f"B{b_idx + 1} Q{q_idx + 1}.{sq_idx + 1}: {(sq['text_uk'] or '')[:40]}"
                )
                question_keys_ordered.append(skey)

    return question_keys_ordered, question_labels


def _get_test_blocks(test_id):
    return Block.find_by_test(test_id) if test_id else Block.find_all()


def _style_header_row(ws, headers):
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = _CELL_ALIGN
    for col_idx, header_val in enumerate(headers, 1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        if header_val in _FIXED_COLS:
            ws.column_dimensions[letter].width = _FIXED_COLS[header_val]
        elif header_val.endswith(" %"):
            ws.column_dimensions[letter].width = 12
        else:
            ws.column_dimensions[letter].width = 18
    ws.freeze_panes = "A2"


def _block_titles_from_subs(subs):
    """Unique block titles across submissions, preserving first-seen order."""
    titles, seen = [], set()
    for s in subs:
        bs = _parse_json_field(s.get("block_scores_json")) or []
        for b in bs if isinstance(bs, list) else []:
            title = b.get("title", f"Block {b.get('id', '?')}")
            if title not in seen:
                seen.add(title)
                titles.append(title)
    return titles


def _summary_row(s, block_titles, question_keys_ordered):
    answers_data = _parse_json_field(s.get("answers_json")) or {}
    if not isinstance(answers_data, dict):
        answers_data = {}

    bs = _parse_json_field(s.get("block_scores_json")) or []
    block_map = {}
    if isinstance(bs, list):
        for b in bs:
            title = b.get("title", f"Block {b.get('id', '?')}")
            block_map[title] = round(b.get("score", 0))

    tg_name = f"{s.get('tg_first_name') or ''} {s.get('tg_last_name') or ''}".strip()
    row = [
        # ── contact / identitate ──
        s["id"], s.get("first_name") or "", s.get("last_name") or "",
        s.get("email") or "", s.get("phone") or "",
        # ── companie ──
        s.get("sector") or "", s.get("company_size") or "",
        s.get("company_age") or "", s.get("company_revenue") or "",
        # ── status / meta ──
        s.get("status", ""), str(s.get("created_at", "")),
        s.get("language", ""), _infer_delivery(s),
        "Так" if s.get("consent") else "Ні",
        # ── Telegram (strâns la un loc) ──
        s.get("tg_username", ""), tg_name, s.get("tg_chat_id", ""),
        # ── scor total ──
        s.get("total_score", ""),
    ]
    # scoruri pe blocuri, apoi răspunsurile la întrebări — la final
    for t in block_titles:
        row.append(block_map.get(t, ""))
    for qkey in question_keys_ordered:
        score = answers_data.get(qkey)
        row.append(score if score is not None else "")
    return row


def _summary_headers(block_titles, question_labels, question_keys_ordered):
    headers = [
        "ID", "Ім'я", "Прізвище", "Email", "Телефон",
        "Сектор", "Розмір компанії", "Вік компанії", "Оборот",
        "Статус", "Дата", "Мова", "Доставка", "Згода",
        "Telegram @user", "Telegram ім'я", "Telegram ChatID",
        "Загальний бал %",
    ]
    headers += [f"{t} %" for t in block_titles]
    headers += [question_labels.get(k, k) for k in question_keys_ordered]
    return headers


def _fill_summary_sheet(ws, subs, block_titles, question_keys_ordered, question_labels):
    headers = _summary_headers(block_titles, question_labels, question_keys_ordered)
    ws.append(headers)
    for s in subs:
        ws.append(_summary_row(s, block_titles, question_keys_ordered))
    _style_header_row(ws, headers)


# Trimmed view for non-completed ("У процесі") submissions: only contact data,
# company info and the total score — no per-question answers, no block scores.
_PROCESSED_HEADERS = [
    "ID", "Ім'я", "Прізвище", "Email", "Телефон",
    "Сектор", "Розмір компанії", "Вік компанії", "Оборот",
    "Статус", "Дата", "Загальний бал %",
]


def _processed_row(s):
    return [
        s["id"], s.get("first_name") or "", s.get("last_name") or "",
        s.get("email") or "", s.get("phone") or "",
        s.get("sector") or "", s.get("company_size") or "",
        s.get("company_age") or "", s.get("company_revenue") or "",
        s.get("status", ""), str(s.get("created_at", "")),
        s.get("total_score", ""),
    ]


def _fill_processed_sheet(ws, subs):
    """Fill a sheet with the trimmed contact+company+score view (no answers)."""
    ws.append(_PROCESSED_HEADERS)
    for s in subs:
        ws.append(_processed_row(s))
    _style_header_row(ws, _PROCESSED_HEADERS)


def _fill_single_user_sheet(ws, sub, question_keys_ordered, question_labels):
    """Render one user's full detail on a dedicated sheet."""
    ws.append(["BizCheck — Індивідуальний звіт"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([])

    name = f"{sub.get('first_name') or ''} {sub.get('last_name') or ''}".strip() or "—"
    pairs = [
        ("Повне ім'я", name),
        ("Email", sub.get("email") or "—"),
        ("Телефон", sub.get("phone") or "—"),
        ("Сектор", sub.get("sector") or "—"),
        ("Розмір компанії", sub.get("company_size") or "—"),
        ("Вік компанії", sub.get("company_age") or "—"),
        ("Оборот", sub.get("company_revenue") or "—"),
        ("Мова", sub.get("language") or "—"),
        ("Обраний канал доставки", _infer_delivery(sub)),
        ("Telegram @username", sub.get("tg_username") or "—"),
        ("Telegram відображуване ім'я", f"{sub.get('tg_first_name') or ''} {sub.get('tg_last_name') or ''}".strip() or "—"),
        ("Статус", sub.get("status") or "—"),
        ("Дата заповнення", str(sub.get("created_at") or "—")),
        ("Загальний бал %", sub.get("total_score", "—")),
    ]
    for k, v in pairs:
        ws.append([k, v])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    ws.append([])
    ws.append(["Бали за блоками"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

    bs = _parse_json_field(sub.get("block_scores_json")) or []
    if isinstance(bs, list):
        for b in bs:
            ws.append([b.get("title", "—"), f"{round(b.get('score', 0))}%"])

    ws.append([])
    ws.append(["Детальні відповіді за питаннями"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append(["Питання", "Бал"])
    for c in ws[ws.max_row]:
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL

    answers = _parse_json_field(sub.get("answers_json")) or {}
    if not isinstance(answers, dict):
        answers = {}
    for qkey in question_keys_ordered:
        ws.append([question_labels.get(qkey, qkey), answers.get(qkey, "")])

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = _CELL_ALIGN
    ws.column_dimensions['A'].width = 46
    ws.column_dimensions['B'].width = 56


def _filter_submissions_for_test(test_id):
    subs = Submission.find_all(test_id=test_id)
    return subs


def _created_date(created_at):
    """Best-effort extraction of a ``date`` from a submission's ``created_at``.

    Accepts a ``datetime``, a ``date``, or an ISO string. Returns ``None`` when
    the value cannot be parsed (caller decides how to treat those rows).
    """
    if created_at is None:
        return None
    if isinstance(created_at, datetime):
        return created_at.date()
    if isinstance(created_at, date):
        return created_at
    try:
        return date.fromisoformat(str(created_at)[:10])
    except (ValueError, TypeError):
        return None


# ──────────────────────────────────────────────────────────────────
# Public builders
# ──────────────────────────────────────────────────────────────────

def build_test_combined_workbook(test_id, date_from=None, date_to=None):
    """Multi-sheet workbook: summary sheets + one sheet per completed user.

    Summary sheets, in order:
      • "Зведення"  — toți participanții (general), cu răspunsuri la întrebări
      • "Завершені" — doar cei care au terminat tot (status == completed)
      • "У процесі" — cei nefinalizați (started / in_progress / abandoned),
                       doar date de contact + companie + scor total (fără răspunsuri)
    Then one detail sheet per *completed* user (cei în proces nu primesc foaie
    individuală cu răspunsuri — apar doar pe foaia trimisă "У процесі").

    When ``date_from`` and/or ``date_to`` (``datetime.date``) are provided,
    submissions are kept only when their ``created_at`` (as a date) falls in the
    inclusive range. Rows whose ``created_at`` can't be parsed are INCLUDED
    (kept) so that filtering never silently drops data.
    """
    subs = _filter_submissions_for_test(test_id)

    if date_from is not None or date_to is not None:
        def _in_range(s):
            d = _created_date(s.get("created_at"))
            if d is None:
                return True  # unparseable → keep, don't lose data
            if date_from is not None and d < date_from:
                return False
            if date_to is not None and d > date_to:
                return False
            return True
        subs = [s for s in subs if _in_range(s)]

    blocks = _get_test_blocks(test_id)
    question_keys_ordered, question_labels = _collect_questions_for_blocks(blocks)
    block_titles = _block_titles_from_subs(subs)

    completed = [s for s in subs if (s.get("status") or "") == "completed"]
    in_progress = [s for s in subs if (s.get("status") or "") != "completed"]

    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "Зведення"
    _fill_summary_sheet(summary, subs, block_titles, question_keys_ordered, question_labels)

    finished_ws = wb.create_sheet(title="Завершені")
    _fill_summary_sheet(finished_ws, completed, block_titles, question_keys_ordered, question_labels)

    in_progress_ws = wb.create_sheet(title="У процесі")
    _fill_processed_sheet(in_progress_ws, in_progress)

    # One sheet per completed user. Name: "{id}_FirstLast" truncated to 31 chars.
    # Over MAX_USER_SHEETS completed users we omit the individual sheets to avoid
    # OOM / request timeouts — the complete data still lives in the summary sheets
    # ("Зведення" / "Завершені" / "У процесі") and in the admin panel.
    used_names = {"Зведення", "Завершені", "У процесі"}
    for s in completed[:MAX_USER_SHEETS]:
        raw = f"{s['id']}_{s.get('first_name') or ''}_{s.get('last_name') or ''}"
        base = _safe_sheet_name(raw)
        name, i = base, 1
        while name in used_names:
            suffix = f"_{i}"
            name = base[: 31 - len(suffix)] + suffix
            i += 1
        used_names.add(name)
        ws = wb.create_sheet(title=name)
        _fill_single_user_sheet(ws, s, question_keys_ordered, question_labels)

    return wb


def build_all_submissions_workbook():
    """Global multi-sheet workbook across ALL tests (admin "Submissions" tab).

    Same layout as the per-test combined workbook, minus per-user sheets
    (the global table can hold thousands of rows):
      • "Зведення"  — toate submisiile, cu răspunsuri la întrebări
      • "Завершені" — status == completed, cu răspunsuri
      • "У процесі" — nefinalizați, doar contact + companie + scor total
    """
    subs = Submission.find_all()
    blocks = Block.find_all()
    question_keys_ordered, question_labels = _collect_questions_for_blocks(blocks)
    block_titles = _block_titles_from_subs(subs)

    completed = [s for s in subs if (s.get("status") or "") == "completed"]
    in_progress = [s for s in subs if (s.get("status") or "") != "completed"]

    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "Зведення"
    _fill_summary_sheet(summary, subs, block_titles, question_keys_ordered, question_labels)

    finished_ws = wb.create_sheet(title="Завершені")
    _fill_summary_sheet(finished_ws, completed, block_titles, question_keys_ordered, question_labels)

    in_progress_ws = wb.create_sheet(title="У процесі")
    _fill_processed_sheet(in_progress_ws, in_progress)

    return wb


def build_single_user_workbook(submission_id):
    sub = Submission.find_by_id(submission_id)
    if not sub:
        raise ValueError("Submission not found")
    blocks = _get_test_blocks(sub.get("test_id"))
    question_keys_ordered, question_labels = _collect_questions_for_blocks(blocks)

    wb = openpyxl.Workbook()
    ws = wb.active
    name = f"{sub.get('first_name') or ''} {sub.get('last_name') or ''}".strip() or f"sub{submission_id}"
    ws.title = _safe_sheet_name(name)
    _fill_single_user_sheet(ws, sub, question_keys_ordered, question_labels)
    return wb, _safe_filename(f"{name}_{sub.get('created_at')}")


def _fmt_score(val) -> str:
    """Render a numeric score as 'NN%'; '—' when missing/non-numeric."""
    if val is None or val == "":
        return "—"
    try:
        return f"{round(float(val))}%"
    except (TypeError, ValueError):
        return "—"


def build_single_user_report_html(submission_id) -> tuple[str, str]:
    """Render one submission's full report as a standalone, printable HTML page.

    Works for ANY submission regardless of whether a PDF was uploaded — the
    admin can always review the report from the data we store. All dynamic
    values are HTML-escaped (defense-in-depth on top of write-path sanitization).

    Returns (html_string, display_name).
    """
    sub = Submission.find_by_id(submission_id)
    if not sub:
        raise ValueError("Submission not found")
    blocks = _get_test_blocks(sub.get("test_id"))
    question_keys_ordered, question_labels = _collect_questions_for_blocks(blocks)

    e = html.escape
    name = f"{sub.get('first_name') or ''} {sub.get('last_name') or ''}".strip() or "—"

    info_pairs = [
        ("Повне ім'я", name),
        ("Email", sub.get("email") or "—"),
        ("Телефон", sub.get("phone") or "—"),
        ("Сектор", sub.get("sector") or "—"),
        ("Розмір компанії", sub.get("company_size") or "—"),
        ("Вік компанії", sub.get("company_age") or "—"),
        ("Оборот", sub.get("company_revenue") or "—"),
        ("Мова", sub.get("language") or "—"),
        ("Канал доставки", _infer_delivery(sub)),
        ("Telegram @username", sub.get("tg_username") or "—"),
        ("Статус", sub.get("status") or "—"),
        ("Дата заповнення", str(sub.get("created_at") or "—")),
    ]
    info_rows = "".join(
        f"<tr><th>{e(k)}</th><td>{e(str(v))}</td></tr>" for k, v in info_pairs
    )

    bs = _parse_json_field(sub.get("block_scores_json")) or []
    block_rows = ""
    if isinstance(bs, list):
        for b in bs:
            if not isinstance(b, dict):
                continue
            title = e(str(b.get("title", "—")))
            block_rows += f"<tr><td>{title}</td><td class='num'>{_fmt_score(b.get('score'))}</td></tr>"
    if not block_rows:
        block_rows = "<tr><td colspan='2' class='empty'>Немає балів за блоками</td></tr>"

    answers = _parse_json_field(sub.get("answers_json")) or {}
    if not isinstance(answers, dict):
        answers = {}
    q_rows = ""
    for qkey in question_keys_ordered:
        val = answers.get(qkey)
        val_str = "" if val is None else e(str(val))
        q_rows += f"<tr><td>{e(question_labels.get(qkey, qkey))}</td><td class='num'>{val_str}</td></tr>"
    if not q_rows:
        q_rows = "<tr><td colspan='2' class='empty'>Немає записаних відповідей</td></tr>"

    page = f"""<!doctype html>
<html lang="uk">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>BizCheck — Звіт {e(name)}</title>
<style>
  :root {{ --ink:#1d2530; --muted:#6b7685; --line:#e3e8ef; --head:#0b2e4f; --accent:#0b6; }}
  * {{ box-sizing:border-box; }}
  body {{ font-family:-apple-system,Segoe UI,Roboto,Arial,sans-serif; color:var(--ink);
         margin:0; padding:32px; background:#f5f7fa; }}
  .sheet {{ max-width:860px; margin:0 auto; background:#fff; border:1px solid var(--line);
            border-radius:10px; padding:32px 36px; }}
  h1 {{ font-size:22px; margin:0 0 4px; color:var(--head); }}
  .sub {{ color:var(--muted); font-size:13px; margin-bottom:24px; }}
  h2 {{ font-size:15px; margin:28px 0 10px; color:var(--head);
        border-bottom:2px solid var(--line); padding-bottom:6px; }}
  table {{ width:100%; border-collapse:collapse; font-size:13px; }}
  th, td {{ text-align:left; padding:8px 10px; border-bottom:1px solid var(--line);
           vertical-align:top; }}
  .info th {{ width:200px; color:var(--muted); font-weight:600; }}
  td.num {{ text-align:right; width:90px; font-variant-numeric:tabular-nums; font-weight:600; }}
  .empty {{ color:var(--muted); font-style:italic; text-align:center; }}
  .total {{ display:inline-block; margin-top:10px; padding:6px 14px; background:var(--head);
            color:#fff; border-radius:6px; font-weight:700; font-size:15px; }}
  @media print {{ body {{ background:#fff; padding:0; }} .sheet {{ border:0; }} }}
</style>
</head>
<body>
  <div class="sheet">
    <h1>BizCheck — Індивідуальний звіт</h1>
    <div class="sub">ID подання #{e(str(sub.get('id', '')))}</div>

    <h2>Дані компанії та контакт</h2>
    <table class="info">{info_rows}</table>

    <h2>Загальний бал</h2>
    <span class="total">{_fmt_score(sub.get('total_score'))}</span>

    <h2>Бали за блоками</h2>
    <table>{block_rows}</table>

    <h2>Детальні відповіді за питаннями</h2>
    <table>{q_rows}</table>
  </div>
</body>
</html>"""
    return page, name


def workbook_to_bytes(wb) -> bytes:
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def build_pdfs_zip_for_test(test_id, max_bytes=None) -> str:
    """ZIP every submission's stored PDF for this test, written to a temp file.

    Streams to a NamedTemporaryFile on disk (not BytesIO) to keep memory flat.
    When ``max_bytes`` is a number, the on-disk size is checked after each PDF
    and ``ExportTooLarge`` is raised once it would be exceeded. When ``max_bytes``
    is ``None`` (default), NO size limit is applied — streaming from disk avoids
    OOM regardless of archive size.

    Returns the filesystem PATH to the temp .zip (caller is responsible for
    removing it once served).
    """
    subs = _filter_submissions_for_test(test_id)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
    try:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zf:
            for s in subs:
                pdf = Submission.get_pdf(s["id"])
                if not pdf:
                    continue
                date_part = str(s.get("created_at") or "").replace(":", "-").replace(" ", "_")[:19]
                contact = s.get("phone") or s.get("email") or ""
                stem = _safe_filename(
                    f"{s['id']}_{s.get('first_name') or ''}_{s.get('last_name') or ''}_{contact}_{date_part}"
                )
                zf.writestr(f"{stem}.pdf", pdf)
                if max_bytes is not None:
                    tmp.flush()
                    if os.fstat(tmp.fileno()).st_size > max_bytes:
                        raise ExportTooLarge()
    except BaseException:
        tmp.close()
        try:
            os.remove(tmp.name)
        except OSError:
            pass
        raise
    tmp.close()
    return tmp.name


def build_excels_zip_for_test(test_id) -> bytes:
    """ZIP containing one Excel per submission (individual detailed report)."""
    subs = _filter_submissions_for_test(test_id)
    blocks = _get_test_blocks(test_id)
    question_keys_ordered, question_labels = _collect_questions_for_blocks(blocks)

    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for s in subs:
            wb = openpyxl.Workbook()
            ws = wb.active
            name = f"{s.get('first_name') or ''} {s.get('last_name') or ''}".strip() or f"sub{s['id']}"
            ws.title = _safe_sheet_name(name)
            _fill_single_user_sheet(ws, s, question_keys_ordered, question_labels)

            date_part = str(s.get("created_at") or "").replace(":", "-").replace(" ", "_")[:19]
            stem = _safe_filename(f"{s['id']}_{name}_{date_part}")
            zf.writestr(f"{stem}.xlsx", workbook_to_bytes(wb))
    buf.seek(0)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────
# Bot picker / single-file helpers
# ──────────────────────────────────────────────────────────────────

def list_submissions_for_picker(test_id, q=None, limit=20):
    """Most-recent submissions for a test, shaped for the bot's inline picker.

    Returns at most ``limit`` dicts {"id", "label"}, newest first, optionally
    filtered by ``q`` (case-insensitive match in first/last name or @username).
    label = "First Last" → else "@username" → else "#<id>"; a trailing
    "· …NNNN" (last 4 phone digits) is appended when a phone is on file.
    """
    subs = Submission.find_all(test_id=test_id) or []
    subs = sorted(subs, key=lambda s: str(s.get("created_at") or ""), reverse=True)

    if q:
        needle = q.strip().lower()
        if needle:
            subs = [
                s for s in subs
                if needle in (s.get("first_name") or "").lower()
                or needle in (s.get("last_name") or "").lower()
                or needle in (s.get("tg_username") or "").lower()
            ]

    out = []
    for s in subs[:limit]:
        name = f"{s.get('first_name') or ''} {s.get('last_name') or ''}".strip()
        if name:
            label = name
        elif s.get("tg_username"):
            label = f"@{s['tg_username'].lstrip('@')}"
        else:
            label = f"#{s['id']}"
        phone = s.get("phone") or ""
        digits = re.sub(r"\D", "", phone)
        if digits:
            label = f"{label} · …{digits[-4:]}"
        out.append({"id": s["id"], "label": label})
    return out


def single_pdf_filename(sub) -> str:
    """Filename for a single submission's stored PDF, based on contact data."""
    contact = sub.get("phone") or sub.get("email") or ""
    stem = _safe_filename(
        f"{sub['id']}_{sub.get('first_name') or ''}_{sub.get('last_name') or ''}_{contact}"
    )
    return stem + ".pdf"


def export_basename(test_id, kind) -> str:
    """Base filename (no extension) for a test export: Виписка_<kind>_<test>_<date>."""
    test = Test.find_by_id(test_id) or {}
    name = test.get("name_uk") or test.get("name_en") or f"Test {test_id}"
    return f"Виписка_{kind}_{_safe_filename(name)}_{date.today().isoformat()}"
