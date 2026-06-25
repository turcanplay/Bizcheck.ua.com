"""Submission routes — public quiz submission + admin data access."""
import re
import base64
import json
import logging

log = logging.getLogger(__name__)
from flask import Blueprint, request, jsonify, Response
from services.submission_service import (
    create_submission, update_submission, save_submission_pdf,
    get_all_submissions, get_submission_pdf, get_submission_detail,
    delete_submission, delete_all_submissions,
)
from models.submission import Submission
from middleware.admin_middleware import admin_required, submission_owner_or_admin
from utils.validators import clean_text, clean_optional

submissions_bp = Blueprint("submissions", __name__, url_prefix="/api_crowe_bizcheck/submissions")

# Allowed fields for PATCH /submissions/{id} (token-gated).
# Anything not listed here is silently dropped — even with a valid token, the
# client can only mutate the operational/PII fields it owns. tg_* and pdf_data
# stay off-limits (bot-only / dedicated endpoints).
_PATCH_ALLOWED = {
    "sector", "company_size", "company_age", "company_revenue", "status",
    "language", "answers_json", "block_scores_json",
    "selected_answers_json", "total_score",
    "first_name", "last_name", "email", "phone", "consent",
}

_VALID_STATUSES = {"started", "in_progress", "completed", "abandoned"}

_EMAIL_RE = re.compile(r'^[^@\s]{1,64}@[^@\s]{1,253}\.[^@\s]{1,63}$')

_PHONE_RE = re.compile(r'^\+?[\d\s\-()]{7,20}$')

def _strip(val, max_len=200):
    """HTML-stripping, control-char-removing string normalizer.

    All user-submitted free-text passes through here so a payload like
    `<svg/onload=...>` becomes the empty string before it ever hits the DB.
    Defense in depth: React already escapes on render, but storing a clean
    value protects every future consumer (PDF generator, exports, etc.).
    """
    return clean_text(val, max_len=max_len)


@submissions_bp.route("", methods=["POST"])
def create():
    data = request.get_json(silent=True) or {}

    errors = []
    first_name = _strip(data.get("first_name"), 30)
    last_name = _strip(data.get("last_name"), 30)
    email = _strip(data.get("email"), 254).lower()
    phone = _strip(data.get("phone"), 20)

    # Contact info is now collected after the quiz; only validate fields if present.
    if email and not _EMAIL_RE.match(email):
        errors.append("Invalid email address")
    if phone and not _PHONE_RE.match(phone):
        errors.append("Invalid phone number")
    if errors:
        return jsonify({"errors": errors}), 400

    # Coerce test_id to int — a malformed value (string, list, …) reaching the
    # DB layer raised HTTP 500 (input-validation finding). Reject cleanly.
    test_id = data.get("test_id")
    if test_id is not None:
        try:
            test_id = int(test_id)
        except (TypeError, ValueError):
            return jsonify({"errors": ["Invalid test_id"]}), 400
    test_slug = _strip(data.get("test_slug"), 64).lower() or None

    try:
        sub = create_submission(
            first_name or None, last_name or None, email or None, phone or None,
            bool(data.get("consent", False)),
            test_id=test_id, test_slug=test_slug,
        )
        # `submission_token` is returned ONLY by create (RETURNING _SELECT_COLS_WITH_TOKEN);
        # update_submission below uses _SELECT_COLS which excludes the token by design.
        # Preserve it across the language patch so the client receives it exactly once.
        created_token = sub.get("submission_token") if sub else None
        # language: whitelist strict — coloana e VARCHAR(5) și doar ro/ru sunt valide.
        # Fără asta, un input cu caractere speciale e expandat de bleach (ex. " → &quot;),
        # depășește VARCHAR(5) și PostgreSQL aruncă „value too long" → HTTP 500.
        lang_in = data.get("language")
        language = lang_in if lang_in in ("uk", "en") else "uk"
        if sub and language:
            updated = update_submission(sub["id"], {"language": language})
            if updated:
                updated["submission_token"] = created_token
                sub = updated
        return jsonify({"submission": sub}), 201
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


_PII_WRITE_ONCE = ("first_name", "last_name", "email", "phone", "consent")


@submissions_bp.route("/<int:sub_id>", methods=["PATCH"])
@submission_owner_or_admin
def update(sub_id):
    data = request.get_json(silent=True) or {}
    filtered = {k: v for k, v in data.items() if k in _PATCH_ALLOWED}
    # Sanitize free-text fields; max_len e fixat la dimensiunea coloanei din DB ca să
    # evităm overflow de VARCHAR (bleach poate expanda caracterele → „value too long" → 500).
    for str_field, mx in (("sector", 200), ("company_size", 50), ("company_age", 50), ("company_revenue", 50), ("status", 20)):
        if str_field in filtered and isinstance(filtered[str_field], str):
            filtered[str_field] = clean_text(filtered[str_field], max_len=mx)
    # language: whitelist strict (coloana VARCHAR(5); doar ro/ru).
    if "language" in filtered:
        filtered["language"] = filtered["language"] if filtered["language"] in ("uk", "en") else "uk"
    if "first_name" in filtered and isinstance(filtered["first_name"], str):
        filtered["first_name"] = clean_optional(filtered["first_name"], max_len=30)
    if "last_name" in filtered and isinstance(filtered["last_name"], str):
        filtered["last_name"] = clean_optional(filtered["last_name"], max_len=30)
    if "email" in filtered and isinstance(filtered["email"], str):
        em = clean_text(filtered["email"], max_len=254).lower()
        if em and not _EMAIL_RE.match(em):
            return jsonify({"error": "Invalid email address"}), 400
        filtered["email"] = em or None
    if "phone" in filtered and isinstance(filtered["phone"], str):
        ph = clean_text(filtered["phone"], max_len=20)
        if ph and not _PHONE_RE.match(ph):
            return jsonify({"error": "Invalid phone number"}), 400
        filtered["phone"] = ph or None
    if "consent" in filtered:
        filtered["consent"] = bool(filtered["consent"])
    if "status" in filtered and filtered["status"] not in _VALID_STATUSES:
        return jsonify({"error": "Invalid status"}), 400
    if "total_score" in filtered:
        try:
            filtered["total_score"] = max(0, min(100, float(filtered["total_score"])))
        except (TypeError, ValueError):
            del filtered["total_score"]

    # Write-once guard: PII fields can only be set if not already present.
    # Protects against a third party with a known sub_id overwriting
    # someone else's contact info.
    pii_fields_in_patch = [f for f in _PII_WRITE_ONCE if f in filtered]
    if pii_fields_in_patch:
        try:
            existing = get_submission_detail(sub_id)
        except ValueError as e:
            return jsonify({"error": str(e)}), 404
        for f in pii_fields_in_patch:
            current = existing.get(f)
            # `consent` is a bool: False is "not yet given", True is locked-in.
            already_set = (current is True) if f == "consent" else bool(current)
            if already_set:
                del filtered[f]

    try:
        sub = update_submission(sub_id, filtered)
    except ValueError as e:
        return jsonify({"error": str(e)}), 404

    # Notify the sales team once the lead has a name + contact channel.
    # Fire-and-forget + fire-once (claimed atomically inside the service).
    try:
        from services.sales_notify import maybe_notify_sales
        maybe_notify_sales(sub_id)
    except Exception:
        pass  # notification must never break the user-facing save
    return jsonify({"submission": sub})


@submissions_bp.route("/<int:sub_id>/pdf", methods=["POST"])
@submission_owner_or_admin
def upload_pdf(sub_id):
    data = request.get_json(silent=True) or {}
    pdf_base64 = data.get("pdf")
    if not pdf_base64:
        return jsonify({"error": "PDF data is required"}), 400
    if len(pdf_base64) > 28_000_000:
        return jsonify({"error": "PDF too large (max 20 MB)"}), 413
    try:
        pdf_bytes = base64.b64decode(pdf_base64, validate=True)
    except Exception:
        return jsonify({"error": "Invalid PDF data"}), 400
    if not pdf_bytes.startswith(b"%PDF"):
        return jsonify({"error": "File is not a valid PDF"}), 400
    try:
        save_submission_pdf(sub_id, pdf_bytes)
        return jsonify({"message": "PDF saved"})
    except ValueError as e:
        return jsonify({"error": str(e)}), 404


@submissions_bp.route("/<int:sub_id>/send-email", methods=["POST"])
@submission_owner_or_admin
def send_email(sub_id):
    """Dispatch the stored report to the submission's email address.

    Fire-and-forget: queues a background send and returns 202 immediately.
    All the heavy lifting (validation, localized template, token-gated download
    link) lives in services.report_email so the Telegram flow reuses it verbatim.
    """
    from services.report_email import dispatch_report_email
    ok, reason = dispatch_report_email(sub_id)
    if ok:
        return jsonify({"message": "Email queued for delivery"}), 202

    status_map = {
        "not_found":     (404, "Submission not found"),
        "no_email":      (400, "No valid email on submission"),
        "pdf_not_ready": (409, "Report PDF not ready yet. Please retry in a moment."),
        "error":         (500, "Failed to queue email"),
    }
    code, msg = status_map.get(reason, (500, "Failed to queue email"))
    log.warning("[send-email] sub %s not sent: %s", sub_id, reason)
    return jsonify({"error": msg}), code


@submissions_bp.route("/<int:sub_id>/report.pdf", methods=["GET"])
def public_report_pdf(sub_id):
    """Public, token-gated report PDF — the target of the email download button.

    The emailed link carries ?t=<submission_token>. We validate the token
    against the id manually (no cookie/CSRF — the link must open in any browser,
    even one the user isn't logged into). Wrong/missing token → 403, identical
    to an unknown id, so the id space can't be enumerated.

    Served `inline` so the browser OPENS the report in a PDF viewer tab; the
    viewer's own toolbar still lets the user download or print it.
    """
    token = request.args.get("t", "")
    if not Submission.find_id_by_token(sub_id, token):
        return jsonify({"error": "Forbidden"}), 403

    try:
        pdf_bytes = get_submission_pdf(sub_id)
    except ValueError:
        pdf_bytes = None
    if not pdf_bytes:
        return jsonify({"error": "Report not available"}), 404

    try:
        sub = get_submission_detail(sub_id)
        first = re.sub(r'[^\w\-]', '_', (sub.get("first_name") or "").strip())[:30]
        last = re.sub(r'[^\w\-]', '_', (sub.get("last_name") or "").strip())[:30]
        fname = f"BizCheck_{first}_{last}.pdf".replace("__", "_").strip("_")
    except Exception:
        fname = ""
    if not fname or fname == "BizCheck_.pdf":
        fname = f"BizCheck_{sub_id}.pdf"

    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{fname}"'},
    )


@submissions_bp.route("", methods=["GET"])
@admin_required
def get_all():
    test_id = request.args.get("test_id", type=int)
    subs = get_all_submissions(test_id=test_id)
    return jsonify({"submissions": subs, "count": len(subs)})


@submissions_bp.route("/<int:sub_id>", methods=["GET"])
@admin_required
def get_detail(sub_id):
    try:
        sub = get_submission_detail(sub_id)
        return jsonify({"submission": sub})
    except ValueError as e:
        return jsonify({"error": str(e)}), 404


@submissions_bp.route("/<int:sub_id>/pdf", methods=["GET"])
@admin_required
def download_pdf(sub_id):
    try:
        sub = get_submission_detail(sub_id)
        pdf_bytes = get_submission_pdf(sub_id)
        first = re.sub(r'[^\w\-]', '_', (sub.get("first_name") or "").strip())[:30]
        last = re.sub(r'[^\w\-]', '_', (sub.get("last_name") or "").strip())[:30]
        fname = f"BizCheck_{first}_{last}.pdf" if first or last else f"BizCheck_report_{sub_id}.pdf"
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except ValueError as e:
        return jsonify({"error": str(e)}), 404


@submissions_bp.route("/<int:sub_id>", methods=["DELETE"])
@admin_required
def delete_one(sub_id):
    try:
        delete_submission(sub_id)
        return jsonify({"message": "Submission deleted"})
    except ValueError as e:
        return jsonify({"error": str(e)}), 404


@submissions_bp.route("", methods=["DELETE"])
@admin_required
def delete_all():
    delete_all_submissions()
    return jsonify({"message": "All submissions deleted"})


@submissions_bp.route("/export/excel", methods=["GET"])
@admin_required
def export_excel():
    """GET /api/submissions/export/excel — Export all submissions as Excel (admin)."""
    from io import BytesIO
    import openpyxl
    from models.question import Question
    from models.block import Block

    subs = get_all_submissions()
    blocks = Block.find_all()
    all_questions = []
    question_labels = {}
    question_keys_ordered = []

    for b_idx, b in enumerate(blocks):
        block_questions = Question.find_by_block(b["id"])
        top_level = [q for q in block_questions if not q.get("parent_question_id")]
        for q_idx, q in enumerate(top_level):
            key = f"b{b['id']}q{q['id']}"
            label = f"B{b_idx+1} Q{q_idx+1}: {(q['text_uk'] or '')[:40]}"
            question_labels[key] = label
            question_keys_ordered.append(key)
            all_questions.append(q)
            subs_qs = [sq for sq in block_questions if sq.get("parent_question_id") == q["id"]]
            for sq_idx, sq in enumerate(subs_qs):
                skey = f"b{b['id']}q{sq['id']}"
                slabel = f"B{b_idx+1} Q{q_idx+1}.{sq_idx+1}: {(sq['text_uk'] or '')[:40]}"
                question_labels[skey] = slabel
                question_keys_ordered.append(skey)

    from models.answer import Answer
    answer_text_map = {}
    for q in all_questions:
        answers = Answer.find_by_question(q["id"])
        for a in answers:
            if q["id"] not in answer_text_map:
                answer_text_map[q["id"]] = {}
            answer_text_map[q["id"]][f"a{a['id']}"] = a["text_uk"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Submissions"

    block_titles = []
    block_title_set = set()
    for s in subs:
        bs = s.get("block_scores_json")
        if bs and isinstance(bs, list):
            for b in bs:
                title = b.get("title", f"Block {b.get('id', '?')}")
                if title not in block_title_set:
                    block_title_set.add(title)
                    block_titles.append(title)

    headers = ["ID", "First Name", "Last Name", "Email", "Phone", "Sector",
               "Company Size", "Company Age", "Sales Revenue", "Total Score %"]
    headers += [f"{t} %" for t in block_titles]
    headers += ["Status", "Consent", "Date", "TG Chat ID", "TG Username", "TG First Name", "TG Last Name"]
    headers += [question_labels.get(k, k) for k in question_keys_ordered]
    ws.append(headers)

    for s in subs:
        row = [
            s["id"], s["first_name"], s["last_name"],
            s.get("email", ""), s.get("phone", ""),
            s.get("sector", ""), s.get("company_size", ""), s.get("company_age", ""),
            s.get("company_revenue", ""),
            s.get("total_score", ""),
        ]
        bs = s.get("block_scores_json")
        block_map = {}
        if bs and isinstance(bs, list):
            for b in bs:
                title = b.get("title", f"Block {b.get('id', '?')}")
                block_map[title] = round(b.get("score", 0))
        for t in block_titles:
            row.append(block_map.get(t, ""))
        row += [
            s.get("status", ""),
            "Yes" if s.get("consent") else "No",
            s["created_at"],
            s.get("tg_chat_id", ""),
            s.get("tg_username", ""),
            s.get("tg_first_name", ""),
            s.get("tg_last_name", ""),
        ]
        answers_data = s.get("answers_json")
        if answers_data and isinstance(answers_data, str):
            try:
                answers_data = json.loads(answers_data)
            except (json.JSONDecodeError, TypeError):
                answers_data = {}
        if not isinstance(answers_data, dict):
            answers_data = {}
        for qkey in question_keys_ordered:
            score = answers_data.get(qkey)
            row.append(score if score is not None else "")
        ws.append(row)

    from openpyxl.styles import Alignment, Font, PatternFill
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell_align = Alignment(wrap_text=True, vertical="top")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = cell_align

    fixed_cols = {
        "ID": 5, "First Name": 12, "Last Name": 12, "Email": 20,
        "Phone": 14, "Sector": 14, "Company Size": 10, "Company Age": 10,
        "Total Score %": 10, "Status": 9, "Consent": 8, "Date": 16,
        "TG Chat ID": 14, "TG Username": 16, "TG First Name": 14, "TG Last Name": 14,
    }
    for col_idx, header_val in enumerate(headers, 1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        if header_val in fixed_cols:
            ws.column_dimensions[letter].width = fixed_cols[header_val]
        elif header_val.endswith(" %"):
            ws.column_dimensions[letter].width = 12
        else:
            ws.column_dimensions[letter].width = 14
    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bizcheck_submissions.xlsx"},
    )


# ─────────────────────────────────────────────────────────────
#  Per-test report exports (used by the admin "Rapoarte" tab)
# ─────────────────────────────────────────────────────────────

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@submissions_bp.route("/<int:sub_id>/export/excel", methods=["GET"])
@admin_required
def export_single_user_excel(sub_id):
    """Detailed single-user Excel (one sheet with company info + answers)."""
    from services.export_service import build_single_user_workbook, workbook_to_bytes
    try:
        wb, stem = build_single_user_workbook(sub_id)
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    return Response(
        workbook_to_bytes(wb),
        mimetype=_XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="BizCheck_{stem}.xlsx"'},
    )


@submissions_bp.route("/tests/<int:test_id>/export/excel-combined", methods=["GET"])
@admin_required
def export_test_combined_excel(test_id):
    """One Excel file for a test — summary sheet + one detail sheet per user."""
    from services.export_service import build_test_combined_workbook, workbook_to_bytes
    wb = build_test_combined_workbook(test_id)
    return Response(
        workbook_to_bytes(wb),
        mimetype=_XLSX_MIME,
        headers={
            "Content-Disposition": f'attachment; filename="BizCheck_test_{test_id}_combined.xlsx"',
        },
    )


@submissions_bp.route("/tests/<int:test_id>/export/excels-zip", methods=["GET"])
@admin_required
def export_test_excels_zip(test_id):
    """ZIP of per-user Excel files for a test."""
    from services.export_service import build_excels_zip_for_test
    data = build_excels_zip_for_test(test_id)
    return Response(
        data,
        mimetype="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="BizCheck_test_{test_id}_excels.zip"',
        },
    )


@submissions_bp.route("/tests/<int:test_id>/export/pdfs-zip", methods=["GET"])
@admin_required
def export_test_pdfs_zip(test_id):
    """ZIP of per-user PDF files for a test."""
    from services.export_service import build_pdfs_zip_for_test
    data = build_pdfs_zip_for_test(test_id)
    return Response(
        data,
        mimetype="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="BizCheck_test_{test_id}_pdfs.zip"',
        },
    )
