"""Unit tests for services.export_service — Excel/ZIP/HTML export builders.

Pure helpers are tested directly; the workbook builders are tested with the
model layer monkeypatched (no DB).
"""
from datetime import date, datetime

import openpyxl
import pytest

from services import export_service as ex


# ---------------------------------------------------------------------------
# Pure string / parsing helpers
# ---------------------------------------------------------------------------
class TestSafeNames:
    def test_safe_filename_replaces_bad_chars(self):
        assert ex._safe_filename("a/b:c*?") == "a_b_c_"

    def test_safe_filename_empty_defaults(self):
        assert ex._safe_filename("") == "report"
        assert ex._safe_filename(None) == "report"

    def test_safe_filename_truncates_to_80(self):
        assert len(ex._safe_filename("x" * 200)) == 80

    def test_safe_sheet_name_strips_invalid(self):
        assert ex._safe_sheet_name("a/b\\c?[d]") == "a_b_c__d_"

    def test_safe_sheet_name_caps_at_31(self):
        assert len(ex._safe_sheet_name("y" * 50)) == 31

    def test_safe_sheet_name_empty_defaults(self):
        assert ex._safe_sheet_name("   ") == "Sheet"


class TestInferDelivery:
    def test_telegram_wins(self):
        assert ex._infer_delivery({"tg_chat_id": 1, "email": "a@b"}) == "Telegram"

    def test_email_second(self):
        assert ex._infer_delivery({"email": "a@b"}) == "Email"

    def test_pdf_fallback(self):
        assert ex._infer_delivery({}) == "PDF"


class TestParseJsonField:
    def test_parses_json_string(self):
        assert ex._parse_json_field('{"a": 1}') == {"a": 1}

    def test_passthrough_dict(self):
        assert ex._parse_json_field({"a": 1}) == {"a": 1}

    def test_bad_json_returns_none(self):
        assert ex._parse_json_field("{not json") is None


class TestFmtScore:
    def test_rounds_numeric(self):
        assert ex._fmt_score(82.4) == "82%"
        assert ex._fmt_score("90") == "90%"

    def test_missing_is_dash(self):
        assert ex._fmt_score(None) == "—"
        assert ex._fmt_score("") == "—"

    def test_non_numeric_is_dash(self):
        assert ex._fmt_score("abc") == "—"


class TestCreatedDate:
    def test_from_datetime(self):
        assert ex._created_date(datetime(2026, 1, 2, 3, 4)) == date(2026, 1, 2)

    def test_from_date(self):
        assert ex._created_date(date(2026, 1, 2)) == date(2026, 1, 2)

    def test_from_iso_string(self):
        assert ex._created_date("2026-01-02T10:00:00") == date(2026, 1, 2)

    def test_none_returns_none(self):
        assert ex._created_date(None) is None

    def test_unparseable_returns_none(self):
        assert ex._created_date("garbage") is None


class TestBlockTitlesFromSubs:
    def test_unique_preserving_first_seen_order(self):
        subs = [
            {"block_scores_json": '[{"title": "A"}, {"title": "B"}]'},
            {"block_scores_json": '[{"title": "B"}, {"title": "C"}]'},
        ]
        assert ex._block_titles_from_subs(subs) == ["A", "B", "C"]

    def test_handles_missing_and_bad_json(self):
        subs = [{"block_scores_json": None}, {"block_scores_json": "bad"}]
        assert ex._block_titles_from_subs(subs) == []


class TestSummaryRow:
    def test_row_maps_block_scores_and_answers(self):
        sub = {
            "id": 1, "first_name": "Ion", "last_name": "P", "email": "a@b.md",
            "phone": "+373", "sector": "IT", "company_size": "10", "company_age": "3",
            "company_revenue": "1M", "language": "ro", "consent": True,
            "total_score": 80, "status": "completed", "created_at": "2026-01-01",
            "block_scores_json": '[{"title": "Bloc A", "score": 75.6}]',
            "answers_json": '{"b1q1": 3}',
        }
        row = ex._summary_row(sub, ["Bloc A"], ["b1q1"])
        assert row[0] == 1                    # id
        assert row[1] == "Ion"                # prenume first
        assert row[-1] == 3                   # answer value last
        assert 76 in row                      # block score rounded
        assert "Da" in row                    # consent rendered (RO)

    def test_header_and_row_lengths_match(self):
        # Guard against the headers/row drifting out of sync on future edits.
        headers = ex._summary_headers(["Bloc A"], {"b1q1": "Q1"}, ["b1q1"])
        row = ex._summary_row({"id": 1}, ["Bloc A"], ["b1q1"])
        assert len(headers) == len(row)

    def test_contact_first_answers_last(self):
        headers = ex._summary_headers([], {}, ["b1q1"])
        assert headers[:5] == ["ID", "Prenume", "Nume", "Email", "Telefon"]
        assert headers[-1] == "Q1" or headers[-1] == "b1q1"   # question at the end


# ---------------------------------------------------------------------------
# list_submissions_for_picker — bot inline picker shaping
# ---------------------------------------------------------------------------
class TestPicker:
    def _patch(self, monkeypatch, subs):
        monkeypatch.setattr("models.submission.Submission.find_all",
                            staticmethod(lambda test_id=None: subs))

    def test_name_label_and_phone_suffix(self, monkeypatch):
        self._patch(monkeypatch, [
            {"id": 1, "first_name": "Ion", "last_name": "Pop", "phone": "079123456",
             "created_at": "2026-01-02"},
        ])
        out = ex.list_submissions_for_picker(1)
        assert out[0]["id"] == 1
        assert out[0]["label"].startswith("Ion Pop")
        assert "…3456" in out[0]["label"]

    def test_username_fallback(self, monkeypatch):
        self._patch(monkeypatch, [
            {"id": 2, "first_name": "", "last_name": "", "tg_username": "@ion",
             "created_at": "2026-01-02"},
        ])
        out = ex.list_submissions_for_picker(1)
        assert out[0]["label"] == "@ion"

    def test_id_fallback(self, monkeypatch):
        self._patch(monkeypatch, [
            {"id": 3, "first_name": "", "last_name": "", "created_at": "2026-01-02"},
        ])
        out = ex.list_submissions_for_picker(1)
        assert out[0]["label"] == "#3"

    def test_search_filter_and_limit(self, monkeypatch):
        subs = [{"id": i, "first_name": f"User{i}", "last_name": "",
                 "created_at": f"2026-01-{i:02d}"} for i in range(1, 30)]
        subs.append({"id": 99, "first_name": "Zebra", "last_name": "", "created_at": "2026-02-01"})
        self._patch(monkeypatch, subs)
        out = ex.list_submissions_for_picker(1, q="zebra")
        assert len(out) == 1 and out[0]["id"] == 99

    def test_limit_caps_results(self, monkeypatch):
        subs = [{"id": i, "first_name": f"U{i}", "last_name": "",
                 "created_at": f"2026-01-{i:02d}"} for i in range(1, 40)]
        self._patch(monkeypatch, subs)
        out = ex.list_submissions_for_picker(1, limit=5)
        assert len(out) == 5


class TestSinglePdfFilename:
    def test_builds_from_contact(self):
        sub = {"id": 7, "first_name": "Ana", "last_name": "P", "phone": "079"}
        name = ex.single_pdf_filename(sub)
        assert name.endswith(".pdf")
        assert "7" in name and "Ana" in name


# ---------------------------------------------------------------------------
# Workbook builders (models monkeypatched)
# ---------------------------------------------------------------------------
_SUB_COMPLETED = {
    "id": 1, "first_name": "Ion", "last_name": "Pop", "email": "a@b.md",
    "phone": "079", "sector": "IT", "status": "completed", "consent": True,
    "total_score": 80, "created_at": "2026-01-01", "test_id": 1,
    "block_scores_json": '[{"title": "Bloc A", "score": 80}]',
    "answers_json": '{"b1q1": 3}',
}
_SUB_INPROGRESS = {
    "id": 2, "first_name": "Ana", "last_name": "", "status": "in_progress",
    "total_score": None, "created_at": "2026-01-02", "test_id": 1,
}


@pytest.fixture
def patch_models(monkeypatch):
    monkeypatch.setattr("models.submission.Submission.find_all",
                        staticmethod(lambda test_id=None: [_SUB_COMPLETED, _SUB_INPROGRESS]))
    monkeypatch.setattr("models.block.Block.find_by_test",
                        staticmethod(lambda t: [{"id": 1}]))
    monkeypatch.setattr("models.block.Block.find_all",
                        staticmethod(lambda: [{"id": 1}]))
    monkeypatch.setattr("models.question.Question.find_by_blocks",
                        staticmethod(lambda ids: [
                            {"id": 1, "block_id": 1, "text_ro": "Întrebare 1",
                             "parent_question_id": None}]))


class TestCombinedWorkbook:
    def test_has_three_summary_sheets_plus_user_sheet(self, patch_models):
        wb = ex.build_test_combined_workbook(1)
        assert wb.sheetnames[:3] == ["Sumar", "Finalizați", "În proces"]
        # one completed user → exactly one extra detail sheet
        assert len(wb.sheetnames) == 4

    def test_date_filter_excludes_out_of_range(self, patch_models):
        wb = ex.build_test_combined_workbook(1, date_from=date(2026, 1, 2))
        # only the in-progress sub (2026-01-02) survives → no completed → no user sheet
        assert "Finalizați" in wb.sheetnames
        assert len(wb.sheetnames) == 3

    def test_workbook_to_bytes_is_valid_xlsx(self, patch_models):
        wb = ex.build_test_combined_workbook(1)
        data = ex.workbook_to_bytes(wb)
        assert data[:2] == b"PK"              # xlsx == zip container


class TestSingleUserHtml:
    def test_report_escapes_and_includes_score(self, monkeypatch, patch_models):
        evil = dict(_SUB_COMPLETED)
        evil["first_name"] = "<script>x</script>"
        monkeypatch.setattr("models.submission.Submission.find_by_id",
                            staticmethod(lambda i: evil))
        html, name = ex.build_single_user_report_html(1)
        assert "<script>x</script>" not in html
        assert "80%" in html

    def test_missing_submission_raises(self, monkeypatch):
        monkeypatch.setattr("models.submission.Submission.find_by_id",
                            staticmethod(lambda i: None))
        with pytest.raises(ValueError):
            ex.build_single_user_report_html(999)


class TestPdfZip:
    def test_skips_subs_without_pdf(self, monkeypatch, tmp_path):
        monkeypatch.setattr("models.submission.Submission.find_all",
                            staticmethod(lambda test_id=None: [_SUB_COMPLETED]))
        monkeypatch.setattr("models.submission.Submission.get_pdf",
                            staticmethod(lambda i: None))
        path = ex.build_pdfs_zip_for_test(1)
        import os, zipfile
        try:
            with zipfile.ZipFile(path) as zf:
                assert zf.namelist() == []   # no PDFs → empty archive
        finally:
            os.remove(path)

    def test_too_large_raises(self, monkeypatch):
        monkeypatch.setattr("models.submission.Submission.find_all",
                            staticmethod(lambda test_id=None: [_SUB_COMPLETED]))
        import os as _os
        # Incompressible payload so DEFLATE can't shrink it below max_bytes.
        blob = _os.urandom(10_000)
        monkeypatch.setattr("models.submission.Submission.get_pdf",
                            staticmethod(lambda i: blob))
        with pytest.raises(ex.ExportTooLarge):
            ex.build_pdfs_zip_for_test(1, max_bytes=100)


class TestExportBasename:
    def test_uses_test_name(self, monkeypatch):
        monkeypatch.setattr("models.test.Test.find_by_id",
                            staticmethod(lambda i: {"name_ro": "BizCheck Pro"}))
        base = ex.export_basename(1, "sumar")
        # _safe_filename keeps spaces (regex [^\w\- ]+), so "BizCheck Pro" survives.
        assert base.startswith("Extras_sumar_BizCheck Pro_")
